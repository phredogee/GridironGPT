from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from gridiron_gpt.draft.fantasy_ranking_export_service import (
    DRAFT_DAY_FIELDS,
    FULL_ANALYSIS_FIELDS,
    build_rankings_xlsx,
)
from gridiron_gpt.draft.fantasy_ranking_score import FantasyRankingScore


def score(player_id="p1", name="Alpha RB", team="BUF", position="RB"):
    return FantasyRankingScore(
        player_id=player_id,
        player_name=name,
        team=team,
        position=position,
        ranking_score=88.5,
        components={
            "baseline": 80.0,
            "market": 90.0,
            "role": 85.0,
            "cortex": 50.0,
            "availability": 100.0,
        },
        weighted_components={},
        provenance={"market": "2026 ADP"},
    )


def population():
    row = score()
    return SimpleNamespace(
        overall=[row],
        by_position={"QB": [], "RB": [row], "WR": [], "TE": []},
    )


def market_views():
    return {
        "p1": SimpleNamespace(
            position_rank=2,
            tier=1,
            consensus_adp=7.5,
            adp_spread=3.0,
            adp_source_count=2,
            draft_value=6.5,
            source_adps={
                "Fantasy Football Calculator": 9.0,
                "NFL Fantasy": 6.0,
            },
        )
    }


def workbook_headers(data: bytes, sheet="Overall"):
    book = load_workbook(BytesIO(data), read_only=True)
    return [cell.value for cell in next(book[sheet].iter_rows())]


def test_draft_day_export_defaults_to_compact_fields():
    data = build_rankings_xlsx(
        population(),
        bye_week_by_team={"BUF": 7},
        football_notes_by_player_id={"p1": "First-team reps"},
        market_views_by_player_id=market_views(),
    )

    assert workbook_headers(data) == [
        "Rank",
        "Player",
        "Pos",
        "Pos Rank",
        "Tier",
        "Team",
        "Bye",
        "Score",
        "Consensus ADP",
        "Draft Value",
        "Football Notes",
    ]


def test_custom_export_can_remove_market_and_role():
    fields = ("rank", "player", "position", "team", "bye", "score")
    data = build_rankings_xlsx(population(), selected_fields=fields)

    headers = workbook_headers(data)
    assert headers == ["Rank", "Player", "Pos", "Team", "Bye", "Score"]
    assert "Market" not in headers
    assert "Role" not in headers


def test_full_analysis_exposes_all_supported_fields():
    data = build_rankings_xlsx(population(), selected_fields=FULL_ANALYSIS_FIELDS)
    assert workbook_headers(data) == [
        "Rank",
        "Player",
        "Pos",
        "Pos Rank",
        "Tier",
        "Team",
        "Bye",
        "Score",
        "Consensus ADP",
        "FFC ADP",
        "NFL ADP",
        "ADP Spread",
        "ADP Sources",
        "Draft Value",
        "Baseline",
        "Market",
        "Role",
        "Cortex",
        "Availability",
        "Football Notes",
        "Provenance",
    ]


def test_export_writes_bye_and_football_note_values():
    data = build_rankings_xlsx(
        population(),
        selected_fields=DRAFT_DAY_FIELDS,
        bye_week_by_team={"BUF": 7},
        football_notes_by_player_id={"p1": "Goal-line role"},
        market_views_by_player_id=market_views(),
    )
    book = load_workbook(BytesIO(data), read_only=True)
    values = [cell.value for cell in next(book["Overall"].iter_rows(min_row=2, max_row=2))]

    assert values[6] == 7
    assert values[10] == "Goal-line role"


def test_full_analysis_writes_individual_and_consensus_adp_values():
    data = build_rankings_xlsx(
        population(),
        selected_fields=FULL_ANALYSIS_FIELDS,
        market_views_by_player_id=market_views(),
    )
    book = load_workbook(BytesIO(data), read_only=True)
    row = [cell.value for cell in next(book["Overall"].iter_rows(min_row=2, max_row=2))]
    values = dict(zip(workbook_headers(data), row))

    assert values["Pos Rank"] == 2
    assert values["Tier"] == 1
    assert values["Consensus ADP"] == 7.5
    assert values["FFC ADP"] == 9.0
    assert values["NFL ADP"] == 6.0
    assert values["ADP Spread"] == 3.0
    assert values["ADP Sources"] == 2
    assert values["Draft Value"] == 6.5
